import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_exact_table_cards(excel_path, output_path, target_tables=None):
    if not os.path.exists(excel_path):
        print(f"错误：找不到输入文件 {excel_path}")
        return

    print("正在读取原始数据...")
    df_datas = pd.read_excel(excel_path, sheet_name='datas').fillna("")
    df_datas.columns = df_datas.columns.str.strip()

    if target_tables:
        df_datas = df_datas[df_datas['table'].isin(target_tables)]

    # 按 table 分组，构建 {table: [(item, is_key), ...]}
    table_items = {}
    for _, row in df_datas.iterrows():
        tbl = row['table']
        if tbl not in table_items:
            table_items[tbl] = []
        table_items[tbl].append((row['item'], row['key'] == '○'))

    IGNORE_KEYS = {'拠点番号'}

    # 按 key signature 分组（忽略 IGNORE_KEYS 中的项目）
    key_groups = {}
    for tbl, items in table_items.items():
        sig = frozenset(item for item, is_key in items if is_key and item not in IGNORE_KEYS)
        key_groups.setdefault(sig, []).append(tbl)

    wb = Workbook()
    wb.remove(wb.active)

    fill_label_sys = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
    fill_label_tbl = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    fill_label_pk  = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")

    fill_val_sys   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fill_val_tbl   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_val_header= PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")

    fill_pk_row    = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

    font_white = Font(name="游ゴシック", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="游ゴシック", size=10, bold=True)
    font_normal = Font(name="游ゴシック", size=10)
    font_red = Font(name="游ゴシック", size=10, bold=True, color="FF0000")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    start_col = 2
    card_width = 3
    col_gap = 1

    for key_sig, tbl_list in key_groups.items():
        sheet_name = "+".join(sorted(key_sig)) if key_sig else "no_key"
        ws = wb.create_sheet(title=sheet_name)
        ws.views.sheetView[0].showGridLines = True

        has_ignored = any(
            item in IGNORE_KEYS and is_key
            for tbl in tbl_list
            for item, is_key in table_items[tbl]
        )
        if has_ignored:
            ws.sheet_properties.tabColor = "FFC000"

        col_offset = start_col

        for table_name in tbl_list:
            c_label = col_offset
            c_content = col_offset + 1
            c_memo = col_offset + 2

            items = table_items[table_name]

            cell = ws.cell(row=2, column=c_label, value="システム")
            cell.fill = fill_label_sys; cell.font = font_white; cell.alignment = align_center

            ws.merge_cells(start_row=2, start_column=c_content, end_row=2, end_column=c_memo)
            cell_sys_val = ws.cell(row=2, column=c_content, value="ATP")
            cell_sys_val.fill = fill_val_sys; cell_sys_val.font = font_normal; cell_sys_val.alignment = align_left

            cell = ws.cell(row=3, column=c_label, value="テーブル")
            cell.fill = fill_label_tbl; cell.font = font_white; cell.alignment = align_center

            ws.merge_cells(start_row=3, start_column=c_content, end_row=3, end_column=c_memo)
            cell_tbl_val = ws.cell(row=3, column=c_content, value=table_name)
            cell_tbl_val.fill = fill_val_tbl; cell_tbl_val.font = font_bold; cell_tbl_val.alignment = align_left

            cell = ws.cell(row=4, column=c_label, value="主キー")
            cell.fill = fill_label_pk; cell.font = font_white; cell.alignment = align_center

            cell_item = ws.cell(row=4, column=c_content, value="项目名")
            cell_item.fill = fill_val_header; cell_item.font = font_bold; cell_item.alignment = align_center

            cell_memo_title = ws.cell(row=4, column=c_memo, value="備考")
            cell_memo_title.fill = fill_val_header; cell_memo_title.font = font_bold; cell_memo_title.alignment = align_center

            current_row = 5
            for item_val, is_key in items:
                cell_pk_lbl = ws.cell(row=current_row, column=c_label)
                cell_key_val = ws.cell(row=current_row, column=c_content, value=item_val)
                cell_memo_val = ws.cell(row=current_row, column=c_memo)

                if is_key:
                    cell_pk_lbl.value = "PK"
                    cell_pk_lbl.fill = fill_pk_row
                    cell_key_val.fill = fill_pk_row
                    cell_memo_val.fill = fill_pk_row

                if item_val in IGNORE_KEYS:
                    cell_pk_lbl.font = font_red
                    cell_key_val.font = font_red
                else:
                    cell_pk_lbl.font = font_bold
                    cell_key_val.font = font_normal
                cell_pk_lbl.alignment = align_center
                cell_key_val.alignment = align_left

                current_row += 1

            for r in range(2, current_row):
                for c in range(c_label, c_memo + 1):
                    ws.cell(row=r, column=c).border = thin_border

            col_offset += card_width + col_gap

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(output_path)
    print(f"转换成功！已输出至: {output_path}")

# ================= 配置与执行 =================
if __name__ == "__main__":
    INPUT_FILE = "input/real_test_data.xlsx"
    OUTPUT_FILE = "output/exact_output_cards.xlsx"

    TABLES_TO_CONVERT = None  # None 表示处理全部表

    generate_exact_table_cards(INPUT_FILE, OUTPUT_FILE, TABLES_TO_CONVERT)