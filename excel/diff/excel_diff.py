# -*- coding:utf-8 -*-

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def compare_sheet(df_old, df_new, sheet_name):
    """
    对比两个DataFrame并返回结果和统计信息
    """
    print(f"\n  正在对比 Sheet: {sheet_name}")
    print(f"  旧文件形状：{df_old.shape} (行数: {len(df_old)}, 列数: {len(df_old.columns)})")
    print(f"  新文件形状：{df_new.shape} (行数: {len(df_new)}, 列数: {len(df_new.columns)})")
    
    # 打印列名信息和类型
    print(f"  旧文件列名：{list(df_old.columns)}")
    print(f"  新文件列名：{list(df_new.columns)}")
    print(f"  旧文件列名类型：{[type(col).__name__ for col in df_old.columns]}")
    print(f"  新文件列名类型：{[type(col).__name__ for col in df_new.columns]}")
    
    # 将所有列名统一转换为字符串，确保一致性
    df_old.columns = [str(col) for col in df_old.columns]
    df_new.columns = [str(col) for col in df_new.columns]

    # 确定最大行数和列数
    max_rows = max(len(df_old), len(df_new))
    
    # 找出共同的列和独有的列
    common_cols = list(set(df_old.columns) & set(df_new.columns))
    old_only_cols = [str(col) for col in set(df_old.columns) - set(df_new.columns)]
    new_only_cols = [str(col) for col in set(df_new.columns) - set(df_old.columns)]
    
    print(f"  共同列数：{len(common_cols)}")
    if old_only_cols:
        print(f"  仅在旧文件中存在的列：{old_only_cols}")
    if new_only_cols:
        print(f"  仅在新文件中存在的列：{new_only_cols}")

    # 创建结果DataFrame
    results = []
    
    # 记录差异统计
    added_rows = 0
    deleted_rows = 0
    modified_rows = 0
    unchanged_rows = 0
    
    # 逐行对比（基于行号）
    for i in range(max_rows):
        row_result = {'行号': i + 2}  # Excel行号从2开始（因为有表头）
        
        if i < len(df_old) and i < len(df_new):
            # 两行都存在，对比内容
            row_changed = False
            changed_cells = []
            
            # 对比共同列
            for col in common_cols:
                try:
                    val_old = df_old.iloc[i][col]
                    val_new = df_new.iloc[i][col]
                except (KeyError, TypeError) as e:
                    # 如果列名访问失败，尝试使用 loc
                    val_old = df_old.loc[i, col] if i in df_old.index else None
                    val_new = df_new.loc[i, col] if i in df_new.index else None
                
                # 处理NaN值比较
                is_different = False
                if pd.isna(val_old) and pd.isna(val_new):
                    is_different = False
                elif pd.isna(val_old) or pd.isna(val_new):
                    is_different = True
                else:
                    is_different = (val_old != val_new)
                
                if is_different:
                    row_changed = True
                    changed_cells.append(col)
                    row_result[f'{col}(旧)'] = val_old
                    row_result[f'{col}(新)'] = val_new
                else:
                    row_result[col] = val_old
            
            # 标记仅在旧文件中的列
            for col in old_only_cols:
                try:
                    row_result[f'{col}(已删除)'] = df_old.iloc[i][col]
                except (KeyError, TypeError):
                    # 如果仍然失败，尝试通过位置访问
                    try:
                        col_idx = df_old.columns.get_loc(col)
                        row_result[f'{col}(已删除)'] = df_old.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(已删除)'] = None
            
            # 标记仅在新文件中的列
            for col in new_only_cols:
                try:
                    row_result[f'{col}(新增)'] = df_new.iloc[i][col]
                except (KeyError, TypeError):
                    # 如果仍然失败，尝试通过位置访问
                    try:
                        col_idx = df_new.columns.get_loc(col)
                        row_result[f'{col}(新增)'] = df_new.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(新增)'] = None
            
            if row_changed or old_only_cols or new_only_cols:
                row_result['变更状态'] = '已修改'
                modified_rows += 1
                # 确保所有元素都是字符串
                old_cols_str = ', '.join(old_only_cols) if old_only_cols else ""
                new_cols_str = ', '.join(new_only_cols) if new_only_cols else ""
                row_result['变更详情'] = f"修改了 {len(changed_cells)} 个单元格" + \
                                       (f", 删除列: {old_cols_str}" if old_cols_str else "") + \
                                       (f", 新增列: {new_cols_str}" if new_cols_str else "")
            else:
                row_result['变更状态'] = '未变化'
                unchanged_rows += 1
                row_result['变更详情'] = ''
                
        elif i >= len(df_new):
            # 该行仅存在于旧文件中（已删除）
            row_result['变更状态'] = '已删除'
            deleted_rows += 1
            for col in df_old.columns:
                try:
                    row_result[f'{col}(旧)'] = df_old.iloc[i][col]
                except (KeyError, TypeError):
                    # 尝试通过位置访问
                    try:
                        col_idx = df_old.columns.get_loc(col)
                        row_result[f'{col}(旧)'] = df_old.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(旧)'] = None
            row_result['变更详情'] = '此行在新版本中已被删除'
            
        elif i >= len(df_old):
            # 该行仅存在于新文件中（新增）
            row_result['变更状态'] = '新增'
            added_rows += 1
            for col in df_new.columns:
                try:
                    row_result[f'{col}(新)'] = df_new.iloc[i][col]
                except (KeyError, TypeError):
                    # 尝试通过位置访问
                    try:
                        col_idx = df_new.columns.get_loc(col)
                        row_result[f'{col}(新)'] = df_new.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(新)'] = None
            row_result['变更详情'] = '此行在新版本中新增'
        
        results.append(row_result)
    
    # 创建结果DataFrame
    df_results = pd.DataFrame(results)
    
    # 调整列顺序：把关键信息放前面
    priority_cols = ['行号', '变更状态', '变更详情']
    if df_results.empty:
        df_results = pd.DataFrame(columns=priority_cols)
    else:
        other_cols = [col for col in df_results.columns if col not in priority_cols]
        df_results = df_results[priority_cols + other_cols]
    
    stats = {
        'added_rows': added_rows,
        'deleted_rows': deleted_rows,
        'modified_rows': modified_rows,
        'unchanged_rows': unchanged_rows,
        'common_cols': common_cols,
        'old_only_cols': old_only_cols,
        'new_only_cols': new_only_cols
    }
    
    return df_results, stats


def generate_highlighted_excel(all_sheets_data, output_xlsx):
    """
    生成带高亮标记的Excel差异文件
    标记规则：
    - 新增行  → 绿色背景
    - 删除行  → 红色背景
    - 修改行  → 淡黄背景，修改的单元格亮橙色背景
    - 未变化  → 无标记
    """
    wb = Workbook()
    wb.remove(wb.active)

    FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    FILL_NEW_ROW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_DEL_ROW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FILL_MOD_ROW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    FILL_MOD_CELL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    FONT_HEADER = Font(bold=True, color="FFFFFF")
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    all_stats = {}

    for sheet_name, (df_old, df_new, df_results, stats) in all_sheets_data.items():
        all_stats[sheet_name] = stats
        ws = wb.create_sheet(title=sheet_name[:31])

        # Build columns: 变更状态 + 新文件所有列 + 旧文件独有列
        new_cols = list(df_new.columns)
        old_only_cols = [str(c) for c in df_old.columns if c not in df_new.columns]

        # Determine which cells changed per row
        changed_cells_per_row = {}
        for i in range(len(df_results)):
            status = df_results.iloc[i]['变更状态']
            if status == '已修改':
                changed = set()
                if i < len(df_old) and i < len(df_new):
                    for col in df_new.columns:
                        if col in df_old.columns:
                            v_old = df_old.iloc[i][col]
                            v_new = df_new.iloc[i][col]
                            if pd.isna(v_old) and pd.isna(v_new):
                                continue
                            if pd.isna(v_old) or pd.isna(v_new) or v_old != v_new:
                                changed.add(col)
                        else:
                            # New column in new file — always mark as changed
                            changed.add(col)
                changed_cells_per_row[i] = changed

        display_cols = ['变更状态'] + new_cols + [f'{c} (旧版-已删除)' for c in old_only_cols]

        # Write header
        for ci, col_name in enumerate(display_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        # Write data rows
        max_rows = max(len(df_old), len(df_new))
        excel_row = 2

        for i in range(max_rows):
            status = df_results.iloc[i]['变更状态'] if i < len(df_results) else '未变化'

            # Determine row-level fill first
            if status == '新增':
                row_fill = FILL_NEW_ROW
            elif status == '已删除':
                row_fill = FILL_DEL_ROW
            elif status == '已修改':
                row_fill = FILL_MOD_ROW
            else:
                row_fill = None

            # Write status column
            cell = ws.cell(row=excel_row, column=1, value=status)
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            if status == '已删除':
                # Row only in old file — write old values
                for ci, col in enumerate(new_cols, 2):
                    val = df_old.iloc[i][col] if col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.fill = FILL_DEL_ROW
                    cell.border = THIN_BORDER
                for ci, col in enumerate(old_only_cols, 2 + len(new_cols)):
                    val = df_old.iloc[i][col] if col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.fill = FILL_DEL_ROW
                    cell.border = THIN_BORDER
            else:
                # Row from new file
                changed = changed_cells_per_row.get(i, set())
                for ci, col in enumerate(new_cols, 2):
                    val = df_new.iloc[i][col] if i < len(df_new) else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if status == '新增':
                        cell.fill = FILL_NEW_ROW
                    elif status == '已修改':
                        cell.fill = FILL_MOD_CELL if col in changed else FILL_MOD_ROW
                # Old-only columns for non-deleted rows — show old values
                for ci, col in enumerate(old_only_cols, 2 + len(new_cols)):
                    val = df_old.iloc[i][col] if i < len(df_old) and col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if status == '已修改':
                        cell.fill = FILL_MOD_CELL

            excel_row += 1

        # Auto-fit column widths
        for ci in range(1, len(display_cols) + 1):
            max_len = len(str(display_cols[ci - 1]))
            for row_idx in range(2, excel_row):
                cell_val = ws.cell(row=row_idx, column=ci).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)
        ws.column_dimensions['A'].width = 12

    # Add summary sheet
    ws_summary = wb.create_sheet(title='汇总统计', index=0)
    summary_data = [
        ['Sheet名称', '总行数', '新增行', '删除行', '修改行', '未变化'],
    ]
    total_add = total_del = total_mod = total_unch = 0
    for sname, st in all_stats.items():
        summary_data.append([sname, st['added_rows'] + st['deleted_rows'] + st['modified_rows'] + st['unchanged_rows'],
                             st['added_rows'], st['deleted_rows'], st['modified_rows'], st['unchanged_rows']])
        total_add += st['added_rows']
        total_del += st['deleted_rows']
        total_mod += st['modified_rows']
        total_unch += st['unchanged_rows']
    summary_data.append(['合计', total_add + total_del + total_mod + total_unch,
                         total_add, total_del, total_mod, total_unch])

    for ri, row_data in enumerate(summary_data, 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if ri == 1:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_CENTER

    ws_summary.column_dimensions['A'].width = 20
    for ci in range(2, 7):
        ws_summary.column_dimensions[get_column_letter(ci)].width = 12

    wb.save(output_xlsx)
    print(f"\n  ✅ 高亮标记的Excel差异文件已导出至: {output_xlsx}")


def compare_excel(old_file, new_file, output_file):
    """
    对比两个 Excel 文件的所有sheet并生成文本格式的变更报告（按行号对比）
    :param old_file: 旧版本文件路径
    :param new_file: 新版本文件路径
    :param output_file: 导出的差异结果路径（.txt格式）
    """
    # 检查文件是否存在
    if not os.path.exists(old_file):
        print(f"错误：找不到文件 {old_file}")
        return
    if not os.path.exists(new_file):
        print(f"错误：找不到文件 {new_file}")
        return
    
    print(" 正在加载数据...")
    try:
        # 读取所有sheet
        excel_old = pd.ExcelFile(old_file)
        excel_new = pd.ExcelFile(new_file)
        
        old_sheets = set(excel_old.sheet_names)
        new_sheets = set(excel_new.sheet_names)
        
        all_sheets = old_sheets | new_sheets
        only_in_old = old_sheets - new_sheets
        only_in_new = new_sheets - old_sheets
        common_sheets = old_sheets & new_sheets
        
        print(f"\n旧文件sheets：{sorted(old_sheets)}")
        print(f"新文件sheets：{sorted(new_sheets)}")
        
        if only_in_old:
            print(f"\n仅在旧文件中存在的sheets：{sorted(only_in_old)}")
        if only_in_new:
            print(f"仅在新文件中存在的sheets：{sorted(only_in_new)}")
        
    except Exception as e:
        print(f"错误：读取 Excel 文件失败 - {e}")
        return
    
    # 确定输出文件格式
    output_is_xlsx = output_file.endswith('.xlsx')
    if output_is_xlsx:
        txt_output = output_file.rsplit('.', 1)[0] + '.txt'
        xlsx_output = output_file
    else:
        if not output_file.endswith('.txt'):
            output_file = output_file.rsplit('.', 1)[0] + '.txt'
        txt_output = output_file
        xlsx_output = output_file.rsplit('.', 1)[0] + '_highlight.xlsx'

    # 创建输出文本文件
    print("\n 开始对比所有sheets...")
    all_stats = {}
    highlight_data = {}
    report_lines = []
    report_lines.append("=" * 100)
    report_lines.append("Excel 文件对比报告")
    report_lines.append("=" * 100)
    report_lines.append(f"旧文件: {old_file}")
    report_lines.append(f"新文件: {new_file}")
    report_lines.append(f"生成时间: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("=" * 100)
    report_lines.append("")
    
    # 对比共同的sheets
    print(f"\n 将对比 {len(common_sheets)} 个共同的sheets: {sorted(common_sheets)}")
    for sheet_name in sorted(common_sheets):
        try:
            print(f"\n{'='*60}")
            print(f"开始处理 Sheet: {sheet_name}")
            print(f"{'='*60}")
            
            df_old = pd.read_excel(excel_old, sheet_name=sheet_name, header=None)
            df_new = pd.read_excel(excel_new, sheet_name=sheet_name, header=None)
            df_old.columns = [f'列{i+1}' for i in range(len(df_old.columns))]
            df_new.columns = [f'列{i+1}' for i in range(len(df_new.columns))]
            
            print(f"  读取完成 - 旧文件: {len(df_old)}行, 新文件: {len(df_new)}行")
            
            df_results, stats = compare_sheet(df_old, df_new, sheet_name)
            all_stats[sheet_name] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)
            
            print(f"  对比完成 - 新增:{stats['added_rows']}, 删除:{stats['deleted_rows']}, 修改:{stats['modified_rows']}, 未变化:{stats['unchanged_rows']}")
            
            # 将结果写入文本报告
            report_lines.append("=" * 100)
            report_lines.append(f"📊 Sheet: {sheet_name}")
            report_lines.append("=" * 100)
            report_lines.append("")
            
            # 添加统计信息
            report_lines.append("【统计摘要】")
            report_lines.append(f"  总行数: {len(df_results)}")
            report_lines.append(f"  新增行: {stats['added_rows']} 条")
            report_lines.append(f"  删除行: {stats['deleted_rows']} 条")
            report_lines.append(f"  修改行: {stats['modified_rows']} 条")
            report_lines.append(f"  未变化: {stats['unchanged_rows']} 条")
            if stats['old_only_cols']:
                report_lines.append(f"  已删除列: {', '.join([str(c) for c in stats['old_only_cols']])}")
            if stats['new_only_cols']:
                report_lines.append(f"  新增列: {', '.join([str(c) for c in stats['new_only_cols']])}")
            report_lines.append("")
            
            # 添加详细变更信息
            report_lines.append("【详细变更记录】")
            changed_rows = df_results[df_results['变更状态'] != '未变化']
            report_lines.append(f"  发现 {len(changed_rows)} 行有变更")
            
            if len(changed_rows) > 0:
                for idx, row in changed_rows.iterrows():
                    report_lines.append(f"\n{'─'*80}")
                    report_lines.append(f"📍 行 {int(row['行号'])} [{row['变更状态']}]")
                    report_lines.append(f"{'─'*80}")
                    report_lines.append(f"  变更详情: {row['变更详情']}")
                    report_lines.append("")
                    
                    # 显示具体变更的字段，突出显示修改的内容
                    modified_fields = []
                    unchanged_fields = []
                    
                    for col in df_results.columns:
                        if col not in ['行号', '变更状态', '变更详情']:
                            try:
                                val = row[col]
                                if pd.notna(val) and str(val) != '':
                                    col_str = str(col)
                                    if '(旧)' in col_str:
                                        # 这是旧值，找到对应的新值
                                        base_col = col_str.replace('(旧)', '')
                                        new_col = f'{base_col}(新)'
                                        if new_col in df_results.columns:
                                            new_val = row[new_col]
                                            modified_fields.append({
                                                'field': base_col,
                                                'old': val,
                                                'new': new_val if pd.notna(new_val) else '(空)'
                                            })
                                    elif '(新)' in col_str:
                                        # 检查是否已经作为配对处理过
                                        base_col = col_str.replace('(新)', '')
                                        old_col = f'{base_col}(旧)'
                                        if old_col not in [str(c) for c in df_results.columns]:
                                            # 只有新值，没有旧值（新增列）
                                            modified_fields.append({
                                                'field': base_col,
                                                'old': '(不存在)',
                                                'new': val
                                            })
                                    elif '(已删除)' in col_str:
                                        modified_fields.append({
                                            'field': col_str,
                                            'old': val,
                                            'new': '(已删除)'
                                        })
                                    elif '(新增)' in col_str:
                                        modified_fields.append({
                                            'field': col_str,
                                            'old': '(不存在)',
                                            'new': val
                                        })
                                    else:
                                        # 未变化的字段
                                        unchanged_fields.append((col, val))
                            except Exception as e:
                                pass
                    
                    # 先显示修改的字段（重点突出）
                    if modified_fields:
                        report_lines.append("  ✏️ 修改的字段：")
                        for field_info in modified_fields:
                            report_lines.append(f"    • {field_info['field']}:")
                            report_lines.append(f"        旧值: {field_info['old']}")
                            report_lines.append(f"        新值: {field_info['new']}")
                            if str(field_info['old']) != str(field_info['new']):
                                report_lines.append(f"        ⚠️  已修改")
                        report_lines.append("")
                    
                    # 再显示未变化的字段（简化显示）
                    if unchanged_fields and len(unchanged_fields) <= 10:  # 只显示前10个未变化字段
                        report_lines.append("  ✓ 未变化的字段：")
                        for col, val in unchanged_fields[:10]:
                            report_lines.append(f"    • {col}: {val}")
                        if len(unchanged_fields) > 10:
                            report_lines.append(f"    ... 还有 {len(unchanged_fields) - 10} 个未变化字段")
                        report_lines.append("")
            else:
                report_lines.append("  ✓ 无变更")
            
            report_lines.append("")
            report_lines.append("")
            
        except Exception as e:
            import traceback
            error_msg = f"  错误：对比sheet '{sheet_name}' 时出错 - {str(e)}"
            print(error_msg)
            print(f"  详细错误信息：{traceback.format_exc()}")
            report_lines.append(f"\n❌ 错误：对比sheet '{sheet_name}' 时出错 - {str(e)}\n")
            report_lines.append(f"详细堆栈：{traceback.format_exc()}\n")
    
    # 处理仅在旧文件中的sheets
    for sheet_name in sorted(only_in_old):
        try:
            df_old = pd.read_excel(excel_old, sheet_name=sheet_name, header=None)
            df_old.columns = [f'列{i+1}' for i in range(len(df_old.columns))]
            df_new = pd.DataFrame(columns=df_old.columns)
            
            df_results, stats = compare_sheet(df_old, df_new, f"{sheet_name} (已删除)")
            all_stats[f"{sheet_name} (已删除)"] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)
            
            report_lines.append("=" * 100)
            report_lines.append(f"🗑️ Sheet: {sheet_name} (整个Sheet已删除)")
            report_lines.append("=" * 100)
            report_lines.append("")
            report_lines.append("【统计摘要】")
            report_lines.append(f"  删除行数: {stats['deleted_rows']} 条")
            report_lines.append("")
            report_lines.append("【详细内容】")
            for idx, row in df_results.iterrows():
                report_lines.append(f"\n--- 行 {int(row['行号'])} [已删除] ---")
                for col in df_results.columns:
                    if col not in ['行号', '变更状态', '变更详情']:
                        val = row[col]
                        if pd.notna(val) and str(val) != '':
                            report_lines.append(f"  {col}: {val}")
            report_lines.append("")
            report_lines.append("")
            
        except Exception as e:
            print(f"  错误：处理sheet '{sheet_name}' 时出错 - {e}")
            report_lines.append(f"\n❌ 错误：处理sheet '{sheet_name}' 时出错 - {e}\n")
    
    # 处理仅在新文件中的sheets
    for sheet_name in sorted(only_in_new):
        try:
            df_new = pd.read_excel(excel_new, sheet_name=sheet_name, header=None)
            df_new.columns = [f'列{i+1}' for i in range(len(df_new.columns))]
            df_old = pd.DataFrame(columns=df_new.columns)
            
            df_results, stats = compare_sheet(df_old, df_new, f"{sheet_name} (新增)")
            all_stats[f"{sheet_name} (新增)"] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)
            
            report_lines.append("=" * 100)
            report_lines.append(f"➕ Sheet: {sheet_name} (整个Sheet新增)")
            report_lines.append("=" * 100)
            report_lines.append("")
            report_lines.append("【统计摘要】")
            report_lines.append(f"  新增行数: {stats['added_rows']} 条")
            report_lines.append("")
            report_lines.append("【详细内容】")
            for idx, row in df_results.iterrows():
                report_lines.append(f"\n--- 行 {int(row['行号'])} [新增] ---")
                for col in df_results.columns:
                    if col not in ['行号', '变更状态', '变更详情']:
                        val = row[col]
                        if pd.notna(val) and str(val) != '':
                            report_lines.append(f"  {col}: {val}")
            report_lines.append("")
            report_lines.append("")
            
        except Exception as e:
            print(f"  错误：处理sheet '{sheet_name}' 时出错 - {e}")
            report_lines.append(f"\n❌ 错误：处理sheet '{sheet_name}' 时出错 - {e}\n")
    
    # 添加总结部分
    total_added = sum(stats['added_rows'] for stats in all_stats.values())
    total_deleted = sum(stats['deleted_rows'] for stats in all_stats.values())
    total_modified = sum(stats['modified_rows'] for stats in all_stats.values())
    total_unchanged = sum(stats['unchanged_rows'] for stats in all_stats.values())
    
    report_lines.append("=" * 100)
    report_lines.append("📈 总计（所有sheets）")
    report_lines.append("=" * 100)
    report_lines.append(f"  共对比: {len(all_stats)} 个sheet")
    report_lines.append(f"  新增行: {total_added} 条")
    report_lines.append(f"  删除行: {total_deleted} 条")
    report_lines.append(f"  修改行: {total_modified} 条")
    report_lines.append(f"  未变化: {total_unchanged} 条")
    report_lines.append("=" * 100)
    report_lines.append("")
    report_lines.append("报告生成完成！")
    
    # 写入文本报告
    print("\n 正在保存文本报告...")
    with open(txt_output, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    print(f" 文本报告已保存: {txt_output}")

    # 生成高亮标记的Excel差异文件
    if highlight_data:
        generate_highlighted_excel(highlight_data, xlsx_output)

    # 打印控制台摘要
    print("\n" + "="*70 + " 版本变更快速总结 " + "="*70)
    
    total_added = sum(stats['added_rows'] for stats in all_stats.values())
    total_deleted = sum(stats['deleted_rows'] for stats in all_stats.values())
    total_modified = sum(stats['modified_rows'] for stats in all_stats.values())
    total_unchanged = sum(stats['unchanged_rows'] for stats in all_stats.values())
    
    for sheet_name, stats in sorted(all_stats.items()):
        print(f"\n📊 Sheet: {sheet_name}")
        print(f"  新增行：{stats['added_rows']} 条")
        print(f"  删除行：{stats['deleted_rows']} 条")
        print(f"  修改行：{stats['modified_rows']} 条")
        print(f"  未变化：{stats['unchanged_rows']} 条")
        
        if stats['old_only_cols']:
            print(f"  已删除列：{', '.join([str(c) for c in stats['old_only_cols']])}")
        if stats['new_only_cols']:
            print(f"  新增列：{', '.join([str(c) for c in stats['new_only_cols']])}")
    
    print("\n" + "-" * 140)
    print(f"📈 总计（所有sheets）：")
    print(f"  共对比: {len(all_stats)} 个sheet")
    print(f"  新增行：{total_added} 条")
    print(f"  删除行：{total_deleted} 条")
    print(f"  修改行：{total_modified} 条")
    print(f"  未变化：{total_unchanged} 条")
    print("=" * 140)
    print(f"\n✅ 文本差异报告已导出至: {txt_output}")
    print(f"✅ 高亮标记Excel已导出至: {xlsx_output}")
    print(f"📁 共对比了 {len(all_stats)} 个sheet")
    print("=" * 140)

    return all_stats

# --- 使用示例 ---
if __name__ == "__main__":
    # 不需要指定key_column，直接比较两个文件的差异
    compare_excel(
        old_file="./input/test1_v1.xlsx",
        new_file="./input/test1_v2.xlsx",
        output_file="./version_diff_report.txt"
    )
