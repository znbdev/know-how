# -*- coding:utf-8 -*-

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def compare_sheet(df_old, df_new, sheet_name):
    """
    Compare two DataFrames and return results and statistics
    """
    print(f"\n  Comparing Sheet: {sheet_name}")
    print(f"  Old file shape: {df_old.shape} (rows: {len(df_old)}, cols: {len(df_old.columns)})")
    print(f"  New file shape: {df_new.shape} (rows: {len(df_new)}, cols: {len(df_new.columns)})")

    # Print column name info and types
    print(f"  Old file columns: {list(df_old.columns)}")
    print(f"  New file columns: {list(df_new.columns)}")
    print(f"  Old file column types: {[type(col).__name__ for col in df_old.columns]}")
    print(f"  New file column types: {[type(col).__name__ for col in df_new.columns]}")

    # Convert all column names to strings for consistency
    df_old.columns = [str(col) for col in df_old.columns]
    df_new.columns = [str(col) for col in df_new.columns]

    # Determine max rows and columns
    max_rows = max(len(df_old), len(df_new))

    # Find common and unique columns
    common_cols = list(set(df_old.columns) & set(df_new.columns))
    old_only_cols = [str(col) for col in set(df_old.columns) - set(df_new.columns)]
    new_only_cols = [str(col) for col in set(df_new.columns) - set(df_old.columns)]

    print(f"  Common columns: {len(common_cols)}")
    if old_only_cols:
        print(f"  Columns only in old file: {old_only_cols}")
    if new_only_cols:
        print(f"  Columns only in new file: {new_only_cols}")

    # Create results DataFrame
    results = []

    # Track difference statistics
    added_rows = 0
    deleted_rows = 0
    modified_rows = 0
    unchanged_rows = 0

    # Compare row by row (based on row index)
    for i in range(max_rows):
        row_result = {'Row No.': i + 2}  # Excel row starts at 2 (because of header)

        if i < len(df_old) and i < len(df_new):
            # Both rows exist, compare content
            row_changed = False
            changed_cells = []

            # Compare common columns
            for col in common_cols:
                try:
                    val_old = df_old.iloc[i][col]
                    val_new = df_new.iloc[i][col]
                except (KeyError, TypeError) as e:
                    # If column name access fails, try using loc
                    val_old = df_old.loc[i, col] if i in df_old.index else None
                    val_new = df_new.loc[i, col] if i in df_new.index else None

                # Handle NaN comparison
                is_different = False
                if pd.isna(val_old) and pd.isna(val_new):
                    is_different = False
                elif pd.isna(val_old) or pd.isna(val_new):
                    is_different = True
                else:
                    is_different = (val_old != val_new)

                if is_different:
                    row_changed = True
                    changed_cells.append(col)
                    row_result[f'{col}(Old)'] = val_old
                    row_result[f'{col}(New)'] = val_new
                else:
                    row_result[col] = val_old

            # Mark columns only in old file
            for col in old_only_cols:
                try:
                    row_result[f'{col}(Deleted)'] = df_old.iloc[i][col]
                except (KeyError, TypeError):
                    # If still fails, try accessing by position
                    try:
                        col_idx = df_old.columns.get_loc(col)
                        row_result[f'{col}(Deleted)'] = df_old.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(Deleted)'] = None

            # Mark columns only in new file
            for col in new_only_cols:
                try:
                    row_result[f'{col}(Added)'] = df_new.iloc[i][col]
                except (KeyError, TypeError):
                    # If still fails, try accessing by position
                    try:
                        col_idx = df_new.columns.get_loc(col)
                        row_result[f'{col}(Added)'] = df_new.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(Added)'] = None

            if row_changed or old_only_cols or new_only_cols:
                row_result['Change Status'] = 'Modified'
                modified_rows += 1
                # Ensure all elements are strings
                old_cols_str = ', '.join(old_only_cols) if old_only_cols else ""
                new_cols_str = ', '.join(new_only_cols) if new_only_cols else ""
                row_result['Change Details'] = f"Modified {len(changed_cells)} cell(s)" + \
                                               (f", deleted columns: {old_cols_str}" if old_cols_str else "") + \
                                               (f", added columns: {new_cols_str}" if new_cols_str else "")
            else:
                row_result['Change Status'] = 'Unchanged'
                unchanged_rows += 1
                row_result['Change Details'] = ''

        elif i >= len(df_new):
            # Row only exists in old file (deleted)
            row_result['Change Status'] = 'Deleted'
            deleted_rows += 1
            for col in df_old.columns:
                try:
                    row_result[f'{col}(Old)'] = df_old.iloc[i][col]
                except (KeyError, TypeError):
                    # Try accessing by position
                    try:
                        col_idx = df_old.columns.get_loc(col)
                        row_result[f'{col}(Old)'] = df_old.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(Old)'] = None
            row_result['Change Details'] = 'This row has been deleted in the new version'

        elif i >= len(df_old):
            # Row only exists in new file (added)
            row_result['Change Status'] = 'Added'
            added_rows += 1
            for col in df_new.columns:
                try:
                    row_result[f'{col}(New)'] = df_new.iloc[i][col]
                except (KeyError, TypeError):
                    # Try accessing by position
                    try:
                        col_idx = df_new.columns.get_loc(col)
                        row_result[f'{col}(New)'] = df_new.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(New)'] = None
            row_result['Change Details'] = 'This row is newly added in the new version'

        results.append(row_result)

    # Create results DataFrame
    df_results = pd.DataFrame(results)

    # Reorder columns: put key info first
    priority_cols = ['Row No.', 'Change Status', 'Change Details']
    if df_results.empty:
        df_results = pd.DataFrame(columns=priority_cols)
    else:
        other_cols = [col for col in df_results.columns if col not in priority_cols]
        df_results = df_results[priority_cols + other_cols]

    stats = {
        'added_rows': added_rows,
        'deleted_rows': deleted_rows,
        'modified_rows': modified_rows,
        'unchanged_rows': unchanged_rows,
        'common_cols': common_cols,
        'old_only_cols': old_only_cols,
        'new_only_cols': new_only_cols
    }

    return df_results, stats


def generate_highlighted_excel(all_sheets_data, output_xlsx):
    """
    Generate highlighted Excel diff file
    Highlight rules:
    - Added rows   -> Green background
    - Deleted rows -> Red background
    - Modified rows -> Light yellow background, modified cells bright orange background
    - Unchanged rows -> No highlight
    """
    wb = Workbook()
    wb.remove(wb.active)

    FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    FILL_NEW_ROW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_DEL_ROW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FILL_MOD_ROW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    FILL_MOD_CELL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    FONT_HEADER = Font(bold=True, color="FFFFFF")
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    all_stats = {}

    for sheet_name, (df_old, df_new, df_results, stats) in all_sheets_data.items():
        all_stats[sheet_name] = stats
        ws = wb.create_sheet(title=sheet_name[:31])

        # Build columns: Change Status + all new file columns + old-only columns
        new_cols = list(df_new.columns)
        old_only_cols = [str(c) for c in df_old.columns if c not in df_new.columns]

        # Determine which cells changed per row
        changed_cells_per_row = {}
        for i in range(len(df_results)):
            status = df_results.iloc[i]['Change Status']
            if status == 'Modified':
                changed = set()
                if i < len(df_old) and i < len(df_new):
                    for col in df_new.columns:
                        if col in df_old.columns:
                            v_old = df_old.iloc[i][col]
                            v_new = df_new.iloc[i][col]
                            if pd.isna(v_old) and pd.isna(v_new):
                                continue
                            if pd.isna(v_old) or pd.isna(v_new) or v_old != v_new:
                                changed.add(col)
                        else:
                            # New column in new file — always mark as changed
                            changed.add(col)
                changed_cells_per_row[i] = changed

        display_cols = ['Change Status'] + new_cols + [f'{c} (Old-Deleted)' for c in old_only_cols]

        # Write header
        for ci, col_name in enumerate(display_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        # Write data rows
        max_rows = max(len(df_old), len(df_new))
        excel_row = 2

        for i in range(max_rows):
            status = df_results.iloc[i]['Change Status'] if i < len(df_results) else 'Unchanged'

            # Determine row-level fill first
            if status == 'Added':
                row_fill = FILL_NEW_ROW
            elif status == 'Deleted':
                row_fill = FILL_DEL_ROW
            elif status == 'Modified':
                row_fill = FILL_MOD_ROW
            else:
                row_fill = None

            # Write status column
            cell = ws.cell(row=excel_row, column=1, value=status)
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            if status == 'Deleted':
                # Row only in old file — write old values
                for ci, col in enumerate(new_cols, 2):
                    val = df_old.iloc[i][col] if col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.fill = FILL_DEL_ROW
                    cell.border = THIN_BORDER
                for ci, col in enumerate(old_only_cols, 2 + len(new_cols)):
                    val = df_old.iloc[i][col] if col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.fill = FILL_DEL_ROW
                    cell.border = THIN_BORDER
            else:
                # Row from new file
                changed = changed_cells_per_row.get(i, set())
                for ci, col in enumerate(new_cols, 2):
                    val = df_new.iloc[i][col] if i < len(df_new) else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if status == 'Added':
                        cell.fill = FILL_NEW_ROW
                    elif status == 'Modified':
                        cell.fill = FILL_MOD_CELL if col in changed else FILL_MOD_ROW
                # Old-only columns for non-deleted rows — show old values
                for ci, col in enumerate(old_only_cols, 2 + len(new_cols)):
                    val = df_old.iloc[i][col] if i < len(df_old) and col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if status == 'Modified':
                        cell.fill = FILL_MOD_CELL

            excel_row += 1

        # Auto-fit column widths
        for ci in range(1, len(display_cols) + 1):
            max_len = len(str(display_cols[ci - 1]))
            for row_idx in range(2, excel_row):
                cell_val = ws.cell(row=row_idx, column=ci).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)
        ws.column_dimensions['A'].width = 12

    # Add summary sheet
    ws_summary = wb.create_sheet(title='Summary', index=0)
    summary_data = [
        ['Sheet Name', 'Total Rows', 'Added Rows', 'Deleted Rows', 'Modified Rows', 'Unchanged'],
    ]
    total_add = total_del = total_mod = total_unch = 0
    for sname, st in all_stats.items():
        summary_data.append([sname, st['added_rows'] + st['deleted_rows'] + st['modified_rows'] + st['unchanged_rows'],
                             st['added_rows'], st['deleted_rows'], st['modified_rows'], st['unchanged_rows']])
        total_add += st['added_rows']
        total_del += st['deleted_rows']
        total_mod += st['modified_rows']
        total_unch += st['unchanged_rows']
    summary_data.append(['Total', total_add + total_del + total_mod + total_unch,
                         total_add, total_del, total_mod, total_unch])

    for ri, row_data in enumerate(summary_data, 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if ri == 1:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_CENTER

    ws_summary.column_dimensions['A'].width = 20
    for ci in range(2, 7):
        ws_summary.column_dimensions[get_column_letter(ci)].width = 12

    wb.save(output_xlsx)
    print(f"\n  Highlighted Excel diff file saved to: {output_xlsx}")


def compare_excel(old_file, new_file, output_file):
    """
    Compare two Excel files across all sheets and generate a text-format change report (row-by-row comparison)
    :param old_file: Old version file path
    :param new_file: New version file path
    :param output_file: Output diff result path (.txt format)
    """
    # Check if files exist
    if not os.path.exists(old_file):
        print(f"Error: Cannot find file {old_file}")
        return
    if not os.path.exists(new_file):
        print(f"Error: Cannot find file {new_file}")
        return

    print(" Loading data...")
    try:
        # Read all sheets
        excel_old = pd.ExcelFile(old_file)
        excel_new = pd.ExcelFile(new_file)

        old_sheets = set(excel_old.sheet_names)
        new_sheets = set(excel_new.sheet_names)

        all_sheets = old_sheets | new_sheets
        only_in_old = old_sheets - new_sheets
        only_in_new = new_sheets - old_sheets
        common_sheets = old_sheets & new_sheets

        print(f"\nOld file sheets: {sorted(old_sheets)}")
        print(f"New file sheets: {sorted(new_sheets)}")

        if only_in_old:
            print(f"\nSheets only in old file: {sorted(only_in_old)}")
        if only_in_new:
            print(f"Sheets only in new file: {sorted(only_in_new)}")

    except Exception as e:
        print(f"Error: Failed to read Excel file - {e}")
        return

    # Determine output file format
    output_is_xlsx = output_file.endswith('.xlsx')
    if output_is_xlsx:
        txt_output = output_file.rsplit('.', 1)[0] + '.txt'
        xlsx_output = output_file
    else:
        if not output_file.endswith('.txt'):
            output_file = output_file.rsplit('.', 1)[0] + '.txt'
        txt_output = output_file
        xlsx_output = output_file.rsplit('.', 1)[0] + '_highlight.xlsx'

    # Create output text file
    print("\n Starting to compare all sheets...")
    all_stats = {}
    highlight_data = {}
    report_lines = []
    report_lines.append("=" * 100)
    report_lines.append("Excel File Diff Report")
    report_lines.append("=" * 100)
    report_lines.append(f"Old file: {old_file}")
    report_lines.append(f"New file: {new_file}")
    report_lines.append(f"Generated at: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("=" * 100)
    report_lines.append("")

    # Compare common sheets
    print(f"\n Comparing {len(common_sheets)} common sheets: {sorted(common_sheets)}")
    for sheet_name in sorted(common_sheets):
        try:
            print(f"\n{'='*60}")
            print(f"Processing Sheet: {sheet_name}")
            print(f"{'='*60}")

            df_old = pd.read_excel(excel_old, sheet_name=sheet_name, header=None)
            df_new = pd.read_excel(excel_new, sheet_name=sheet_name, header=None)
            df_old.columns = [f'Col_{i+1}' for i in range(len(df_old.columns))]
            df_new.columns = [f'Col_{i+1}' for i in range(len(df_new.columns))]

            print(f"  Read complete - Old file: {len(df_old)} rows, New file: {len(df_new)} rows")

            df_results, stats = compare_sheet(df_old, df_new, sheet_name)
            all_stats[sheet_name] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)

            print(f"  Comparison complete - Added:{stats['added_rows']}, Deleted:{stats['deleted_rows']}, Modified:{stats['modified_rows']}, Unchanged:{stats['unchanged_rows']}")

            # Write results to text report
            report_lines.append("=" * 100)
            report_lines.append(f"Sheet: {sheet_name}")
            report_lines.append("=" * 100)
            report_lines.append("")

            # Add statistics
            report_lines.append("[Summary]")
            report_lines.append(f"  Total Rows: {len(df_results)}")
            report_lines.append(f"  Added Rows: {stats['added_rows']}")
            report_lines.append(f"  Deleted Rows: {stats['deleted_rows']}")
            report_lines.append(f"  Modified Rows: {stats['modified_rows']}")
            report_lines.append(f"  Unchanged: {stats['unchanged_rows']}")
            if stats['old_only_cols']:
                report_lines.append(f"  Deleted columns: {', '.join([str(c) for c in stats['old_only_cols']])}")
            if stats['new_only_cols']:
                report_lines.append(f"  Added columns: {', '.join([str(c) for c in stats['new_only_cols']])}")
            report_lines.append("")

            # Add detailed change info
            report_lines.append("[Detailed Changes]")
            changed_rows = df_results[df_results['Change Status'] != 'Unchanged']
            report_lines.append(f"  Found {len(changed_rows)} rows with changes")

            if len(changed_rows) > 0:
                for idx, row in changed_rows.iterrows():
                    report_lines.append(f"\n{'─'*80}")
                    report_lines.append(f"Row {int(row['Row No.'])} [{row['Change Status']}]")
                    report_lines.append(f"{'─'*80}")
                    report_lines.append(f"  Change Details: {row['Change Details']}")
                    report_lines.append("")

                    # Show specific changed fields, highlight modified content
                    modified_fields = []
                    unchanged_fields = []

                    for col in df_results.columns:
                        if col not in ['Row No.', 'Change Status', 'Change Details']:
                            try:
                                val = row[col]
                                if pd.notna(val) and str(val) != '':
                                    col_str = str(col)
                                    if '(Old)' in col_str:
                                        # This is old value, find corresponding new value
                                        base_col = col_str.replace('(Old)', '')
                                        new_col = f'{base_col}(New)'
                                        if new_col in df_results.columns:
                                            new_val = row[new_col]
                                            modified_fields.append({
                                                'field': base_col,
                                                'old': val,
                                                'new': new_val if pd.notna(new_val) else '(empty)'
                                            })
                                    elif '(New)' in col_str:
                                        # Check if already processed as a pair
                                        base_col = col_str.replace('(New)', '')
                                        old_col = f'{base_col}(Old)'
                                        if old_col not in [str(c) for c in df_results.columns]:
                                            # Only new value, no old value (added column)
                                            modified_fields.append({
                                                'field': base_col,
                                                'old': '(N/A)',
                                                'new': val
                                            })
                                    elif '(Deleted)' in col_str:
                                        modified_fields.append({
                                            'field': col_str,
                                            'old': val,
                                            'new': '(Deleted)'
                                        })
                                    elif '(Added)' in col_str:
                                        modified_fields.append({
                                            'field': col_str,
                                            'old': '(N/A)',
                                            'new': val
                                        })
                                    else:
                                        # Unchanged fields
                                        unchanged_fields.append((col, val))
                            except Exception as e:
                                pass

                    # Show modified fields first (highlighted)
                    if modified_fields:
                        report_lines.append("  Modified fields:")
                        for field_info in modified_fields:
                            report_lines.append(f"    - {field_info['field']}:")
                            report_lines.append(f"        Old value: {field_info['old']}")
                            report_lines.append(f"        New value: {field_info['new']}")
                            if str(field_info['old']) != str(field_info['new']):
                                report_lines.append(f"        [Modified]")
                        report_lines.append("")

                    # Show unchanged fields (simplified)
                    if unchanged_fields and len(unchanged_fields) <= 10:  # Only show first 10 unchanged fields
                        report_lines.append("  Unchanged fields:")
                        for col, val in unchanged_fields[:10]:
                            report_lines.append(f"    - {col}: {val}")
                        if len(unchanged_fields) > 10:
                            report_lines.append(f"    ... and {len(unchanged_fields) - 10} more unchanged fields")
                        report_lines.append("")
            else:
                report_lines.append("  No changes")

            report_lines.append("")
            report_lines.append("")

        except Exception as e:
            import traceback
            error_msg = f"  Error: Failed to compare sheet '{sheet_name}' - {str(e)}"
            print(error_msg)
            print(f"  Error details: {traceback.format_exc()}")
            report_lines.append(f"\nError: Failed to compare sheet '{sheet_name}' - {str(e)}\n")
            report_lines.append(f"Traceback: {traceback.format_exc()}\n")

    # Process sheets only in old file
    for sheet_name in sorted(only_in_old):
        try:
            df_old = pd.read_excel(excel_old, sheet_name=sheet_name, header=None)
            df_old.columns = [f'Col_{i+1}' for i in range(len(df_old.columns))]
            df_new = pd.DataFrame(columns=df_old.columns)

            df_results, stats = compare_sheet(df_old, df_new, f"{sheet_name} (Deleted)")
            all_stats[f"{sheet_name} (Deleted)"] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)

            report_lines.append("=" * 100)
            report_lines.append(f"Sheet: {sheet_name} (Entire sheet deleted)")
            report_lines.append("=" * 100)
            report_lines.append("")
            report_lines.append("[Summary]")
            report_lines.append(f"  Deleted Rows: {stats['deleted_rows']}")
            report_lines.append("")
            report_lines.append("[Details]")
            for idx, row in df_results.iterrows():
                report_lines.append(f"\n--- Row {int(row['Row No.'])} [Deleted] ---")
                for col in df_results.columns:
                    if col not in ['Row No.', 'Change Status', 'Change Details']:
                        val = row[col]
                        if pd.notna(val) and str(val) != '':
                            report_lines.append(f"  {col}: {val}")
            report_lines.append("")
            report_lines.append("")

        except Exception as e:
            print(f"  Error: Failed to process sheet '{sheet_name}' - {e}")
            report_lines.append(f"\nError: Failed to process sheet '{sheet_name}' - {e}\n")

    # Process sheets only in new file
    for sheet_name in sorted(only_in_new):
        try:
            df_new = pd.read_excel(excel_new, sheet_name=sheet_name, header=None)
            df_new.columns = [f'Col_{i+1}' for i in range(len(df_new.columns))]
            df_old = pd.DataFrame(columns=df_new.columns)

            df_results, stats = compare_sheet(df_old, df_new, f"{sheet_name} (Added)")
            all_stats[f"{sheet_name} (Added)"] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)

            report_lines.append("=" * 100)
            report_lines.append(f"Sheet: {sheet_name} (Entire sheet added)")
            report_lines.append("=" * 100)
            report_lines.append("")
            report_lines.append("[Summary]")
            report_lines.append(f"  Added Rows: {stats['added_rows']}")
            report_lines.append("")
            report_lines.append("[Details]")
            for idx, row in df_results.iterrows():
                report_lines.append(f"\n--- Row {int(row['Row No.'])} [Added] ---")
                for col in df_results.columns:
                    if col not in ['Row No.', 'Change Status', 'Change Details']:
                        val = row[col]
                        if pd.notna(val) and str(val) != '':
                            report_lines.append(f"  {col}: {val}")
            report_lines.append("")
            report_lines.append("")

        except Exception as e:
            print(f"  Error: Failed to process sheet '{sheet_name}' - {e}")
            report_lines.append(f"\nError: Failed to process sheet '{sheet_name}' - {e}\n")

    # Add summary section
    total_added = sum(stats['added_rows'] for stats in all_stats.values())
    total_deleted = sum(stats['deleted_rows'] for stats in all_stats.values())
    total_modified = sum(stats['modified_rows'] for stats in all_stats.values())
    total_unchanged = sum(stats['unchanged_rows'] for stats in all_stats.values())

    report_lines.append("=" * 100)
    report_lines.append("Total (all sheets)")
    report_lines.append("=" * 100)
    report_lines.append(f"  Total sheets compared: {len(all_stats)}")
    report_lines.append(f"  Added Rows: {total_added}")
    report_lines.append(f"  Deleted Rows: {total_deleted}")
    report_lines.append(f"  Modified Rows: {total_modified}")
    report_lines.append(f"  Unchanged: {total_unchanged}")
    report_lines.append("=" * 100)
    report_lines.append("")
    report_lines.append("Report generated successfully!")

    # Write text report
    print("\n Saving text report...")
    with open(txt_output, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    print(f" Text report saved: {txt_output}")

    # Generate highlighted Excel diff file
    if highlight_data:
        generate_highlighted_excel(highlight_data, xlsx_output)

    # Print console summary
    print("\n" + "="*70 + " Version Change Summary " + "="*70)

    total_added = sum(stats['added_rows'] for stats in all_stats.values())
    total_deleted = sum(stats['deleted_rows'] for stats in all_stats.values())
    total_modified = sum(stats['modified_rows'] for stats in all_stats.values())
    total_unchanged = sum(stats['unchanged_rows'] for stats in all_stats.values())

    for sheet_name, stats in sorted(all_stats.items()):
        print(f"\nSheet: {sheet_name}")
        print(f"  Added rows: {stats['added_rows']}")
        print(f"  Deleted rows: {stats['deleted_rows']}")
        print(f"  Modified rows: {stats['modified_rows']}")
        print(f"  Unchanged: {stats['unchanged_rows']}")

        if stats['old_only_cols']:
            print(f"  Deleted columns: {', '.join([str(c) for c in stats['old_only_cols']])}")
        if stats['new_only_cols']:
            print(f"  Added columns: {', '.join([str(c) for c in stats['new_only_cols']])}")

    print("\n" + "-" * 140)
    print(f"Total (all sheets):")
    print(f"  Total sheets compared: {len(all_stats)}")
    print(f"  Added rows: {total_added}")
    print(f"  Deleted rows: {total_deleted}")
    print(f"  Modified rows: {total_modified}")
    print(f"  Unchanged: {total_unchanged}")
    print("=" * 140)
    print(f"\nText diff report exported to: {txt_output}")
    print(f"Highlighted Excel exported to: {xlsx_output}")
    print(f"Compared {len(all_stats)} sheet(s) in total")
    print("=" * 140)

    return all_stats

# --- Usage example ---
if __name__ == "__main__":
    # No key_column needed, directly compare differences between two files
    compare_excel(
        old_file="./input/test1_v1.xlsx",
        new_file="./input/test1_v2.xlsx",
        output_file="./version_diff_report.txt"
    )
