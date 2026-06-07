# -*- coding:utf-8 -*-

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def compare_sheet(df_old, df_new, sheet_name):
    """
    2つのDataFrameを比較し、結果と統計情報を返す
    """
    print(f"\n  Sheetを比較中: {sheet_name}")
    print(f"  旧ファイル形状：{df_old.shape} (行数: {len(df_old)}, 列数: {len(df_old.columns)})")
    print(f"  新ファイル形状：{df_new.shape} (行数: {len(df_new)}, 列数: {len(df_new.columns)})")
    
    # 列名情報と型を表示
    print(f"  旧ファイル列名：{list(df_old.columns)}")
    print(f"  新ファイル列名：{list(df_new.columns)}")
    print(f"  旧ファイル列名型：{[type(col).__name__ for col in df_old.columns]}")
    print(f"  新ファイル列名型：{[type(col).__name__ for col in df_new.columns]}")
    
    # すべての列名を文字列に統一し、一貫性を確保
    df_old.columns = [str(col) for col in df_old.columns]
    df_new.columns = [str(col) for col in df_new.columns]

    # 最大行数と最大列数を決定
    max_rows = max(len(df_old), len(df_new))
    
    # 共通列と固有列を特定
    common_cols = list(set(df_old.columns) & set(df_new.columns))
    old_only_cols = [str(col) for col in set(df_old.columns) - set(df_new.columns)]
    new_only_cols = [str(col) for col in set(df_new.columns) - set(df_old.columns)]
    
    print(f"  共通列数：{len(common_cols)}")
    if old_only_cols:
        print(f"  旧ファイルのみに存在する列：{old_only_cols}")
    if new_only_cols:
        print(f"  新ファイルのみに存在する列：{new_only_cols}")

    # 結果DataFrameを作成
    results = []
    
    # 差分統計を記録
    added_rows = 0
    deleted_rows = 0
    modified_rows = 0
    unchanged_rows = 0
    
    # 行ごとに比較（行番号ベース）
    for i in range(max_rows):
        row_result = {'行番号': i + 2}  # Excelの行番号は2から（ヘッダーがあるため）
        
        if i < len(df_old) and i < len(df_new):
            # 両方の行が存在する、内容を比較
            row_changed = False
            changed_cells = []
            
            # 共通列を比較
            for col in common_cols:
                try:
                    val_old = df_old.iloc[i][col]
                    val_new = df_new.iloc[i][col]
                except (KeyError, TypeError) as e:
                    # 列名でのアクセスが失敗した場合、locを試す
                    val_old = df_old.loc[i, col] if i in df_old.index else None
                    val_new = df_new.loc[i, col] if i in df_new.index else None
                
                # NaN値の比較を処理
                is_different = False
                if pd.isna(val_old) and pd.isna(val_new):
                    is_different = False
                elif pd.isna(val_old) or pd.isna(val_new):
                    is_different = True
                else:
                    is_different = (val_old != val_new)
                
                if is_different:
                    row_changed = True
                    changed_cells.append(col)
                    row_result[f'{col}(旧)'] = val_old
                    row_result[f'{col}(新)'] = val_new
                else:
                    row_result[col] = val_old
            
            # 旧ファイルのみに存在する列をマーク
            for col in old_only_cols:
                try:
                    row_result[f'{col}(削除済)'] = df_old.iloc[i][col]
                except (KeyError, TypeError):
                    # それでも失敗する場合、位置でアクセス
                    try:
                        col_idx = df_old.columns.get_loc(col)
                        row_result[f'{col}(削除済)'] = df_old.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(削除済)'] = None
            
            # 新ファイルのみに存在する列をマーク
            for col in new_only_cols:
                try:
                    row_result[f'{col}(新規)'] = df_new.iloc[i][col]
                except (KeyError, TypeError):
                    # それでも失敗する場合、位置でアクセス
                    try:
                        col_idx = df_new.columns.get_loc(col)
                        row_result[f'{col}(新規)'] = df_new.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(新規)'] = None
            
            if row_changed or old_only_cols or new_only_cols:
                row_result['変更状態'] = '変更あり'
                modified_rows += 1
                # すべての要素を文字列に変換
                old_cols_str = ', '.join(old_only_cols) if old_only_cols else ""
                new_cols_str = ', '.join(new_only_cols) if new_only_cols else ""
                row_result['変更詳細'] = f"{len(changed_cells)} 個のセルを変更" + \
                                       (f", 削除列: {old_cols_str}" if old_cols_str else "") + \
                                       (f", 新規列: {new_cols_str}" if new_cols_str else "")
            else:
                row_result['変更状態'] = '変更なし'
                unchanged_rows += 1
                row_result['変更詳細'] = ''
                
        elif i >= len(df_new):
            # この行は旧ファイルのみに存在する（削除済み）
            row_result['変更状態'] = '削除済み'
            deleted_rows += 1
            for col in df_old.columns:
                try:
                    row_result[f'{col}(旧)'] = df_old.iloc[i][col]
                except (KeyError, TypeError):
                    # 位置でアクセスを試す
                    try:
                        col_idx = df_old.columns.get_loc(col)
                        row_result[f'{col}(旧)'] = df_old.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(旧)'] = None
            row_result['変更詳細'] = 'この行は新バージョンで削除されました'
            
        elif i >= len(df_old):
            # この行は新ファイルのみに存在する（新規追加）
            row_result['変更状態'] = '新規追加'
            added_rows += 1
            for col in df_new.columns:
                try:
                    row_result[f'{col}(新)'] = df_new.iloc[i][col]
                except (KeyError, TypeError):
                    # 位置でアクセスを試す
                    try:
                        col_idx = df_new.columns.get_loc(col)
                        row_result[f'{col}(新)'] = df_new.iloc[i, col_idx]
                    except:
                        row_result[f'{col}(新)'] = None
            row_result['変更詳細'] = 'この行は新バージョンで追加されました'
        
        results.append(row_result)
    
    # 結果DataFrameを作成
    df_results = pd.DataFrame(results)
    
    # 列の順序を調整：重要な情報を前に配置
    priority_cols = ['行番号', '変更状態', '変更詳細']
    if df_results.empty:
        df_results = pd.DataFrame(columns=priority_cols)
    else:
        other_cols = [col for col in df_results.columns if col not in priority_cols]
        df_results = df_results[priority_cols + other_cols]
    
    stats = {
        'added_rows': added_rows,
        'deleted_rows': deleted_rows,
        'modified_rows': modified_rows,
        'unchanged_rows': unchanged_rows,
        'common_cols': common_cols,
        'old_only_cols': old_only_cols,
        'new_only_cols': new_only_cols
    }
    
    return df_results, stats


def generate_highlighted_excel(all_sheets_data, output_xlsx):
    """
    ハイライト付きExcel差分ファイルを生成
    マークルール：
    - 新規行  → 緑色背景
    - 削除行  → 赤色背景
    - 変更行  → 薄黄色背景、変更セルは明るい橙色背景
    - 変更なし  → マークなし
    """
    wb = Workbook()
    wb.remove(wb.active)

    FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    FILL_NEW_ROW = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FILL_DEL_ROW = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    FILL_MOD_ROW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    FILL_MOD_CELL = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    FONT_HEADER = Font(bold=True, color="FFFFFF")
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    all_stats = {}

    for sheet_name, (df_old, df_new, df_results, stats) in all_sheets_data.items():
        all_stats[sheet_name] = stats
        ws = wb.create_sheet(title=sheet_name[:31])

        # Build columns: 変更状態 + 新ファイル全列 + 旧ファイル固有列
        new_cols = list(df_new.columns)
        old_only_cols = [str(c) for c in df_old.columns if c not in df_new.columns]

        # Determine which cells changed per row
        changed_cells_per_row = {}
        for i in range(len(df_results)):
            status = df_results.iloc[i]['変更状態']
            if status == '変更あり':
                changed = set()
                if i < len(df_old) and i < len(df_new):
                    for col in df_new.columns:
                        if col in df_old.columns:
                            v_old = df_old.iloc[i][col]
                            v_new = df_new.iloc[i][col]
                            if pd.isna(v_old) and pd.isna(v_new):
                                continue
                            if pd.isna(v_old) or pd.isna(v_new) or v_old != v_new:
                                changed.add(col)
                        else:
                            # New column in new file — always mark as changed
                            changed.add(col)
                changed_cells_per_row[i] = changed

        display_cols = ['変更状態'] + new_cols + [f'{c} (旧版-削除済)' for c in old_only_cols]

        # Write header
        for ci, col_name in enumerate(display_cols, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER

        # Write data rows
        max_rows = max(len(df_old), len(df_new))
        excel_row = 2

        for i in range(max_rows):
            status = df_results.iloc[i]['変更状態'] if i < len(df_results) else '変更なし'

            # Determine row-level fill first
            if status == '新規追加':
                row_fill = FILL_NEW_ROW
            elif status == '削除済み':
                row_fill = FILL_DEL_ROW
            elif status == '変更あり':
                row_fill = FILL_MOD_ROW
            else:
                row_fill = None

            # Write status column
            cell = ws.cell(row=excel_row, column=1, value=status)
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
            if row_fill:
                cell.fill = row_fill

            if status == '削除済み':
                # Row only in old file — write old values
                for ci, col in enumerate(new_cols, 2):
                    val = df_old.iloc[i][col] if col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.fill = FILL_DEL_ROW
                    cell.border = THIN_BORDER
                for ci, col in enumerate(old_only_cols, 2 + len(new_cols)):
                    val = df_old.iloc[i][col] if col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.fill = FILL_DEL_ROW
                    cell.border = THIN_BORDER
            else:
                # Row from new file
                changed = changed_cells_per_row.get(i, set())
                for ci, col in enumerate(new_cols, 2):
                    val = df_new.iloc[i][col] if i < len(df_new) else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if status == '新規追加':
                        cell.fill = FILL_NEW_ROW
                    elif status == '変更あり':
                        cell.fill = FILL_MOD_CELL if col in changed else FILL_MOD_ROW
                # Old-only columns for non-deleted rows — show old values
                for ci, col in enumerate(old_only_cols, 2 + len(new_cols)):
                    val = df_old.iloc[i][col] if i < len(df_old) and col in df_old.columns else None
                    cell = ws.cell(row=excel_row, column=ci, value=val)
                    cell.border = THIN_BORDER
                    if status == '変更あり':
                        cell.fill = FILL_MOD_CELL

            excel_row += 1

        # Auto-fit column widths
        for ci in range(1, len(display_cols) + 1):
            max_len = len(str(display_cols[ci - 1]))
            for row_idx in range(2, excel_row):
                cell_val = ws.cell(row=row_idx, column=ci).value
                if cell_val is not None:
                    max_len = max(max_len, len(str(cell_val)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)
        ws.column_dimensions['A'].width = 12

    # Add summary sheet
    ws_summary = wb.create_sheet(title='集計', index=0)
    summary_data = [
        ['Sheet名', '総行数', '追加行', '削除行', '変更行', '変更なし'],
    ]
    total_add = total_del = total_mod = total_unch = 0
    for sname, st in all_stats.items():
        summary_data.append([sname, st['added_rows'] + st['deleted_rows'] + st['modified_rows'] + st['unchanged_rows'],
                             st['added_rows'], st['deleted_rows'], st['modified_rows'], st['unchanged_rows']])
        total_add += st['added_rows']
        total_del += st['deleted_rows']
        total_mod += st['modified_rows']
        total_unch += st['unchanged_rows']
    summary_data.append(['合計', total_add + total_del + total_mod + total_unch,
                         total_add, total_del, total_mod, total_unch])

    for ri, row_data in enumerate(summary_data, 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=ri, column=ci, value=val)
            cell.border = THIN_BORDER
            if ri == 1:
                cell.fill = FILL_HEADER
                cell.font = FONT_HEADER
                cell.alignment = ALIGN_CENTER

    ws_summary.column_dimensions['A'].width = 20
    for ci in range(2, 7):
        ws_summary.column_dimensions[get_column_letter(ci)].width = 12

    wb.save(output_xlsx)
    print(f"\n  ✅ ハイライト付きExcel差分ファイルを出力しました: {output_xlsx}")


def compare_excel(old_file, new_file, output_file):
    """
    2つのExcelファイルの全シートを比較し、テキスト形式の変更レポートを生成する（行番号ベース）
    :param old_file: 旧バージョンファイルのパス
    :param new_file: 新バージョンファイルのパス
    :param output_file: 出力される差分結果のパス（.txt形式）
    """
    # ファイルの存在確認
    if not os.path.exists(old_file):
        print(f"エラー：ファイルが見つかりません {old_file}")
        return
    if not os.path.exists(new_file):
        print(f"エラー：ファイルが見つかりません {new_file}")
        return
    
    print(" データを読み込み中...")
    try:
        # 全シートを読み込み
        excel_old = pd.ExcelFile(old_file)
        excel_new = pd.ExcelFile(new_file)
        
        old_sheets = set(excel_old.sheet_names)
        new_sheets = set(excel_new.sheet_names)
        
        all_sheets = old_sheets | new_sheets
        only_in_old = old_sheets - new_sheets
        only_in_new = new_sheets - old_sheets
        common_sheets = old_sheets & new_sheets
        
        print(f"\n旧ファイルsheets：{sorted(old_sheets)}")
        print(f"新ファイルsheets：{sorted(new_sheets)}")
        
        if only_in_old:
            print(f"\n旧ファイルのみに存在するsheets：{sorted(only_in_old)}")
        if only_in_new:
            print(f"新ファイルのみに存在するsheets：{sorted(only_in_new)}")
        
    except Exception as e:
        print(f"エラー：Excelファイルの読み込みに失敗しました - {e}")
        return
    
    # 出力ファイルの形式を決定
    output_is_xlsx = output_file.endswith('.xlsx')
    if output_is_xlsx:
        txt_output = output_file.rsplit('.', 1)[0] + '.txt'
        xlsx_output = output_file
    else:
        if not output_file.endswith('.txt'):
            output_file = output_file.rsplit('.', 1)[0] + '.txt'
        txt_output = output_file
        xlsx_output = output_file.rsplit('.', 1)[0] + '_highlight.xlsx'

    # 出力テキストファイルを作成
    print("\n 全シートの比較を開始...")
    all_stats = {}
    highlight_data = {}
    report_lines = []
    report_lines.append("=" * 100)
    report_lines.append("Excel ファイル比較レポート")
    report_lines.append("=" * 100)
    report_lines.append(f"旧ファイル: {old_file}")
    report_lines.append(f"新ファイル: {new_file}")
    report_lines.append(f"生成日時: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append("=" * 100)
    report_lines.append("")
    
    # 共通sheetsを比較
    print(f"\n {len(common_sheets)} 個の共通sheetsを比較: {sorted(common_sheets)}")
    for sheet_name in sorted(common_sheets):
        try:
            print(f"\n{'='*60}")
            print(f"Sheet処理開始: {sheet_name}")
            print(f"{'='*60}")
            
            df_old = pd.read_excel(excel_old, sheet_name=sheet_name, header=None)
            df_new = pd.read_excel(excel_new, sheet_name=sheet_name, header=None)
            df_old.columns = [f'列{i+1}' for i in range(len(df_old.columns))]
            df_new.columns = [f'列{i+1}' for i in range(len(df_new.columns))]
            
            print(f"  読込完了 - 旧ファイル: {len(df_old)}行, 新ファイル: {len(df_new)}行")
            
            df_results, stats = compare_sheet(df_old, df_new, sheet_name)
            all_stats[sheet_name] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)
            
            print(f"  比較完了 - 追加:{stats['added_rows']}, 削除:{stats['deleted_rows']}, 変更:{stats['modified_rows']}, 変更なし:{stats['unchanged_rows']}")
            
            # 結果をテキストレポートに書き込み
            report_lines.append("=" * 100)
            report_lines.append(f"📊 Sheet: {sheet_name}")
            report_lines.append("=" * 100)
            report_lines.append("")
            
            # 統計情報を追加
            report_lines.append("【統計サマリー】")
            report_lines.append(f"  総行数: {len(df_results)}")
            report_lines.append(f"  追加行: {stats['added_rows']} 行")
            report_lines.append(f"  削除行: {stats['deleted_rows']} 行")
            report_lines.append(f"  変更行: {stats['modified_rows']} 行")
            report_lines.append(f"  変更なし: {stats['unchanged_rows']} 行")
            if stats['old_only_cols']:
                report_lines.append(f"  削除列: {', '.join([str(c) for c in stats['old_only_cols']])}")
            if stats['new_only_cols']:
                report_lines.append(f"  新規列: {', '.join([str(c) for c in stats['new_only_cols']])}")
            report_lines.append("")
            
            # 詳細変更情報を追加
            report_lines.append("【変更詳細】")
            changed_rows = df_results[df_results['変更状態'] != '変更なし']
            report_lines.append(f"  {len(changed_rows)} 行に変更があります")
            
            if len(changed_rows) > 0:
                for idx, row in changed_rows.iterrows():
                    report_lines.append(f"\n{'─'*80}")
                    report_lines.append(f"📍 行 {int(row['行番号'])} [{row['変更状態']}]")
                    report_lines.append(f"{'─'*80}")
                    report_lines.append(f"  変更詳細: {row['変更詳細']}")
                    report_lines.append("")
                    
                    # 具体的な変更フィールドを表示、変更内容を強調
                    modified_fields = []
                    unchanged_fields = []
                    
                    for col in df_results.columns:
                        if col not in ['行番号', '変更状態', '変更詳細']:
                            try:
                                val = row[col]
                                if pd.notna(val) and str(val) != '':
                                    col_str = str(col)
                                    if '(旧)' in col_str:
                                        # 旧値の場合、対応する新値を検索
                                        base_col = col_str.replace('(旧)', '')
                                        new_col = f'{base_col}(新)'
                                        if new_col in df_results.columns:
                                            new_val = row[new_col]
                                            modified_fields.append({
                                                'field': base_col,
                                                'old': val,
                                                'new': new_val if pd.notna(new_val) else '(空)'
                                            })
                                    elif '(新)' in col_str:
                                        # 既にペアとして処理済みか確認
                                        base_col = col_str.replace('(新)', '')
                                        old_col = f'{base_col}(旧)'
                                        if old_col not in [str(c) for c in df_results.columns]:
                                            # 新値のみで旧値がない場合（新規列）
                                            modified_fields.append({
                                                'field': base_col,
                                                'old': '(存在しない)',
                                                'new': val
                                            })
                                    elif '(削除済)' in col_str:
                                        modified_fields.append({
                                            'field': col_str,
                                            'old': val,
                                            'new': '(削除済)'
                                        })
                                    elif '(新規)' in col_str:
                                        modified_fields.append({
                                            'field': col_str,
                                            'old': '(存在しない)',
                                            'new': val
                                        })
                                    else:
                                        # 変更なしフィールド
                                        unchanged_fields.append((col, val))
                            except Exception as e:
                                pass
                    
                    # 変更フィールドを先に表示（重点表示）
                    if modified_fields:
                        report_lines.append("  ✏️ 変更フィールド：")
                        for field_info in modified_fields:
                            report_lines.append(f"    • {field_info['field']}:")
                            report_lines.append(f"        旧値: {field_info['old']}")
                            report_lines.append(f"        新値: {field_info['new']}")
                            if str(field_info['old']) != str(field_info['new']):
                                report_lines.append(f"        ⚠️  変更あり")
                        report_lines.append("")
                    
                    # 変更なしフィールドを表示（簡略表示）
                    if unchanged_fields and len(unchanged_fields) <= 10:  # 最初の10個のみ表示
                        report_lines.append("  ✓ 変更なしフィールド：")
                        for col, val in unchanged_fields[:10]:
                            report_lines.append(f"    • {col}: {val}")
                        if len(unchanged_fields) > 10:
                            report_lines.append(f"    ... 他に {len(unchanged_fields) - 10} 個の変更なしフィールド")
                        report_lines.append("")
            else:
                report_lines.append("  ✓ 変更なし")
            
            report_lines.append("")
            report_lines.append("")
            
        except Exception as e:
            import traceback
            error_msg = f"  エラー：シート '{sheet_name}' の比較中にエラーが発生しました - {str(e)}"
            print(error_msg)
            print(f"  詳細エラー情報：{traceback.format_exc()}")
            report_lines.append(f"\n❌ エラー：シート '{sheet_name}' の比較中にエラーが発生しました - {str(e)}\n")
            report_lines.append(f"詳細スタック：{traceback.format_exc()}\n")
    
    # 旧ファイルのみに存在するsheetsを処理
    for sheet_name in sorted(only_in_old):
        try:
            df_old = pd.read_excel(excel_old, sheet_name=sheet_name, header=None)
            df_old.columns = [f'列{i+1}' for i in range(len(df_old.columns))]
            df_new = pd.DataFrame(columns=df_old.columns)
            
            df_results, stats = compare_sheet(df_old, df_new, f"{sheet_name} (削除済み)")
            all_stats[f"{sheet_name} (削除済み)"] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)
            
            report_lines.append("=" * 100)
            report_lines.append(f"🗑️ Sheet: {sheet_name} (シート全体が削除されました)")
            report_lines.append("=" * 100)
            report_lines.append("")
            report_lines.append("【統計サマリー】")
            report_lines.append(f"  削除行数: {stats['deleted_rows']} 行")
            report_lines.append("")
            report_lines.append("【詳細内容】")
            for idx, row in df_results.iterrows():
                report_lines.append(f"\n--- 行 {int(row['行番号'])} [削除済み] ---")
                for col in df_results.columns:
                    if col not in ['行番号', '変更状態', '変更詳細']:
                        val = row[col]
                        if pd.notna(val) and str(val) != '':
                            report_lines.append(f"  {col}: {val}")
            report_lines.append("")
            report_lines.append("")
            
        except Exception as e:
            print(f"  エラー：シート '{sheet_name}' の処理中にエラーが発生しました - {e}")
            report_lines.append(f"\n❌ エラー：シート '{sheet_name}' の処理中にエラーが発生しました - {e}\n")
    
    # 新ファイルのみに存在するsheetsを処理
    for sheet_name in sorted(only_in_new):
        try:
            df_new = pd.read_excel(excel_new, sheet_name=sheet_name, header=None)
            df_new.columns = [f'列{i+1}' for i in range(len(df_new.columns))]
            df_old = pd.DataFrame(columns=df_new.columns)
            
            df_results, stats = compare_sheet(df_old, df_new, f"{sheet_name} (新規追加)")
            all_stats[f"{sheet_name} (新規追加)"] = stats
            highlight_data[sheet_name] = (df_old, df_new, df_results, stats)
            
            report_lines.append("=" * 100)
            report_lines.append(f"➕ Sheet: {sheet_name} (シート全体が新規追加)")
            report_lines.append("=" * 100)
            report_lines.append("")
            report_lines.append("【統計サマリー】")
            report_lines.append(f"  追加行数: {stats['added_rows']} 行")
            report_lines.append("")
            report_lines.append("【詳細内容】")
            for idx, row in df_results.iterrows():
                report_lines.append(f"\n--- 行 {int(row['行番号'])} [新規追加] ---")
                for col in df_results.columns:
                    if col not in ['行番号', '変更状態', '変更詳細']:
                        val = row[col]
                        if pd.notna(val) and str(val) != '':
                            report_lines.append(f"  {col}: {val}")
            report_lines.append("")
            report_lines.append("")
            
        except Exception as e:
            print(f"  エラー：シート '{sheet_name}' の処理中にエラーが発生しました - {e}")
            report_lines.append(f"\n❌ エラー：シート '{sheet_name}' の処理中にエラーが発生しました - {e}\n")
    
    # 集計部分を追加
    total_added = sum(stats['added_rows'] for stats in all_stats.values())
    total_deleted = sum(stats['deleted_rows'] for stats in all_stats.values())
    total_modified = sum(stats['modified_rows'] for stats in all_stats.values())
    total_unchanged = sum(stats['unchanged_rows'] for stats in all_stats.values())
    
    report_lines.append("=" * 100)
    report_lines.append("📈 合計（全シート）")
    report_lines.append("=" * 100)
    report_lines.append(f"  比較: {len(all_stats)} 個のsheet")
    report_lines.append(f"  追加行: {total_added} 行")
    report_lines.append(f"  削除行: {total_deleted} 行")
    report_lines.append(f"  変更行: {total_modified} 行")
    report_lines.append(f"  変更なし: {total_unchanged} 行")
    report_lines.append("=" * 100)
    report_lines.append("")
    report_lines.append("レポート生成完了！")
    
    # テキストレポートを書き込み
    print("\n テキストレポートを保存中...")
    with open(txt_output, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    print(f" テキストレポートを保存しました: {txt_output}")

    # ハイライト付きExcel差分ファイルを生成
    if highlight_data:
        generate_highlighted_excel(highlight_data, xlsx_output)

    # コンソールサマリーを表示
    print("\n" + "="*70 + " バージョン変更サマリー " + "="*70)
    
    total_added = sum(stats['added_rows'] for stats in all_stats.values())
    total_deleted = sum(stats['deleted_rows'] for stats in all_stats.values())
    total_modified = sum(stats['modified_rows'] for stats in all_stats.values())
    total_unchanged = sum(stats['unchanged_rows'] for stats in all_stats.values())
    
    for sheet_name, stats in sorted(all_stats.items()):
        print(f"\n📊 Sheet: {sheet_name}")
        print(f"  追加行：{stats['added_rows']} 行")
        print(f"  削除行：{stats['deleted_rows']} 行")
        print(f"  変更行：{stats['modified_rows']} 行")
        print(f"  変更なし：{stats['unchanged_rows']} 行")
        
        if stats['old_only_cols']:
            print(f"  削除列：{', '.join([str(c) for c in stats['old_only_cols']])}")
        if stats['new_only_cols']:
            print(f"  新規列：{', '.join([str(c) for c in stats['new_only_cols']])}")
    
    print("\n" + "-" * 140)
    print(f"📈 合計（全シート）：")
    print(f"  比較: {len(all_stats)} 個のsheet")
    print(f"  追加行：{total_added} 行")
    print(f"  削除行：{total_deleted} 行")
    print(f"  変更行：{total_modified} 行")
    print(f"  変更なし：{total_unchanged} 行")
    print("=" * 140)
    print(f"\n✅ テキスト差分レポートを出力しました: {txt_output}")
    print(f"✅ ハイライト付きExcelを出力しました: {xlsx_output}")
    print(f"📁 比較したsheet数: {len(all_stats)}")
    print("=" * 140)

    return all_stats

# --- 使用例 ---
if __name__ == "__main__":
    # key_columnの指定は不要で、2つのファイルの差分を直接比較
    compare_excel(
        old_file="./input/test1_v1.xlsx",
        new_file="./input/test1_v2.xlsx",
        output_file="./version_diff_report.txt"
    )
