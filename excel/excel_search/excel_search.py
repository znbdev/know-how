import os
import sys
from typing import List

from openpyxl import load_workbook


class SearchResult:
    def __init__(self, file: str, sheet: str, cell: str, value: str):
        self.file = file
        self.sheet = sheet
        self.cell = cell
        self.value = value


def search_excel_files(directory: str, keyword: str) -> List[SearchResult]:
    results = []
    keyword_lower = keyword.lower()

    for root, _, files in os.walk(directory):
        for fname in files:
            if not (fname.endswith(".xlsx") or fname.endswith(".xls")):
                continue
            fpath = os.path.join(root, fname)
            try:
                wb = load_workbook(fpath, read_only=True, data_only=True)
            except Exception:
                continue

            for sname in wb.sheetnames:
                ws = wb[sname]
                for row in ws.iter_rows():
                    for cell in row:
                        val = cell.value
                        if val is None:
                            continue
                        if isinstance(val, str) and keyword_lower in val.lower():
                            results.append(SearchResult(
                                file=fpath,
                                sheet=sname,
                                cell=cell.coordinate,
                                value=val[:200],
                            ))
                        elif isinstance(val, (int, float)):
                            if keyword_lower in str(val):
                                results.append(SearchResult(
                                    file=fpath,
                                    sheet=sname,
                                    cell=cell.coordinate,
                                    value=str(val),
                                ))

            wb.close()

    return results


def print_results(results: List[SearchResult], keyword: str):
    if not results:
        print("No results found.")
        return

    kw_lower = keyword.lower()
    print(f"\nFound {len(results)} match(es):\n")
    print(f"{'File':<60} {'Sheet':<12} {'Cell':<8}  Value")
    print("-" * 120)
    for r in results:
        value = r.value
        idx = value.lower().find(kw_lower)
        if idx != -1:
            value = value[:idx] + "[" + value[idx:idx+len(keyword)] + "]" + value[idx+len(keyword):]
        print(f"{r.file:<60} {r.sheet:<12} {r.cell:<8}  {value}")


def excel_search():
    if len(sys.argv) == 3:
        path = sys.argv[1]
        keyword = sys.argv[2]
    else:
        path = input("Directory path: ").strip()
        keyword = input("Keyword: ").strip()

    if not path:
        print("Error: directory path is required")
        sys.exit(1)
    if not keyword:
        print("Error: keyword is required")
        sys.exit(1)
    if not os.path.isdir(path):
        print(f"Error: directory not found: {path}")
        sys.exit(1)

    print("Searching...")
    results = search_excel_files(path, keyword)
    print_results(results, keyword)


if __name__ == "__main__":
    excel_search()
